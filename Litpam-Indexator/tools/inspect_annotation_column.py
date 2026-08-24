#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
inspect_annotation_column.py -- confirms the AddAnnotationData (§7.1) data
source (H2590). The stage[4] "Остаток" item asked "нужен двухколоночный
источник термин->аннотация -- вероятно из XLSX-колонки Краткая аннотация" as
an open question. This script dumps every sheet's header row and, for any
sheet with a "Краткая аннотация" column, samples it and counts non-empty
rows -- confirming all four index sheets (Именной/Географ/Предметы и
термины/Флора и фауна) carry a populated headword->gloss column already,
which is the data AddAnnotationData needs; no separate acquisition step.

Запуск:
    python inspect_annotation_column.py [path-to-workbook.xlsx]

_Автор инструмента: Dr. Mārcis Gasūns · создан 25-08-2026 (H2590 rebuild)._
"""

import sys

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

import openpyxl

DEFAULT_PATH = "xls/derived/Указатель_к_Рамаяне_1_2_2026_08_15b.xlsx"


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        print("SHEET:", ws.title, "cols:", header)
        if not any(h and "аннотац" in str(h).lower() for h in header):
            continue
        idx = [i for i, h in enumerate(header) if h and "аннотац" in str(h).lower()][0]
        count = 0
        nonempty = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            count += 1
            v = row[idx] if idx < len(row) else None
            if v not in (None, ""):
                nonempty += 1
                if nonempty <= 15:
                    print("  sample:", repr(row[0]), "->", repr(v))
        print("  total rows:", count, "non-empty annotation:", nonempty)


if __name__ == "__main__":
    main()

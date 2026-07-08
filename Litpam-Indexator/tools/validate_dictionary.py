#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
validate_dictionary.py — pre-flight-валидатор рабочего словаря указателей «Рамаяны».

Аддитивный, СТРОГО read-only препроцессор. Прогоняется по рабочему листу словаря
(`xls/Указатель_к_Рамаяне_*.xlsx`) ДО дорогой стадии `[1]`/`[3]` в InDesign и ловит
классы дефектов, которые сейчас всплывают только во время многочасового прогона:

  1. prose_in_forms    — свободная проза/комментарий в колонке «Что искать…» (вешает
                         обработку — видео 12, слово «оружие»);
  2. blank_row         — пустая строка внутри диапазона данных (стадия `[1]` не
                         отрабатывает — видео 09);
  3. service_leak      — служебная строка (`=VLOOKUP…`, повтор заголовка), просочившаяся
                         в данные;
  4. level_gap         — разрыв уровней термина (напр. 1 → 3 без 2);
  5. trailing_semicolon— хвостовой `;` в ячейке падежных форм;
  6. duplicate_forms   — дубли форм внутри одного термина;
  7. empty_name        — непустые формы при пустом термине (осиротевшая строка).

Чем он ОТЛИЧАЕТСЯ от `_Ram_Tag_explorer/teg_exp.exe`: `teg_exp` сверяет тег ↔ словарь
*между* файлами (тег есть в тексте, но нет в словаре, и наоборот), читает `.xlsx` через
OLE и требует Windows + установленный Excel. Этот скрипт проверяет *внутренние* дефекты
самого листа словаря, кроссплатформенно (openpyxl, без Excel/OLE). Инструменты
дополняют друг друга, а не дублируют.

Ничего в `.xlsx` НЕ меняется — только отчёт; оператор правит вручную (автосанитайзер —
отдельная будущая задача). Единственная внешняя зависимость — `openpyxl`.

Запуск:
    python validate_dictionary.py <путь-к-xlsx>
    python validate_dictionary.py <путь-к-xlsx> --report отчёт.md
    python validate_dictionary.py <путь-к-xlsx> --sheet "Именной"   # только один лист

Коды возврата: 0 — дефектов не найдено; 1 — найдены дефекты; 2 — ошибка запуска.

_Автор инструмента: Dr. Mārcis Gasūns · создан 08-07-2026 (H363)._
"""

import argparse
import re
import sys

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

# --- Рабочие листы (см. Litpam-Indexator/CLAUDE.md). Служебные листы пропускаются. ---
WORKING_SHEETS = ["Именной", "Географ", "Предметы и термины", "Флора и фауна", "Все_Теги"]

# --- Читаем только первыеN колонок: осмысленные поля словаря — в пределах первых 8;
#     дальше идут рабочие FuzzyLookup-колонки (у «Именного» max_col ≈ 1044), их копировать
#     в InDesign оператор не должен, поэтому в проверку они не входят. ---
MAX_COL = 10

# --- Признаки прозы/комментария в колонке форм (вешают обработку). ---
PROSE_KEYWORDS = re.compile(
    r"не\s+искать|без\s+тег|вручну|см\.\s|коммент|примечан|"
    r"NB\b|только\s+с\s+тег|запрещ",
    re.IGNORECASE,
)
# Форма — это 1–5 слов; сегмент из 6+ слов почти наверняка проза.
LONG_SEGMENT_WORDS = 6
FORMULA = re.compile(r"^\s*=|VLOOKUP", re.IGNORECASE)


def norm(v):
    """Ячейка → строка без хвостовых пробелов (None → '')."""
    if v is None:
        return ""
    return str(v).strip()


def find_columns(header):
    """По строке-заголовку вернуть 1-based индексы колонок name/tag/forms (или None)."""
    cols = {"name": None, "tag": None, "forms": None}
    for idx, cell in enumerate(header, start=1):
        h = norm(cell).lower()
        if not h:
            continue
        if cols["name"] is None and (h.startswith("имя") or "теги из текста" in h):
            cols["name"] = idx
        elif cols["tag"] is None and "тег для поиска" in h:
            cols["tag"] = idx
        elif cols["forms"] is None and "что искать" in h:
            cols["forms"] = idx
    return cols


def depth(name):
    """Уровень термина = число сегментов через '\\' (напр. 'Агни\\Анала' → 2)."""
    return name.count("\\") + 1


def split_forms(forms):
    """Падежные формы → список непустых сегментов, разделённых ';'."""
    return [s.strip() for s in forms.split(";") if s.strip()]


def check_sheet(name, rows):
    """rows: список кортежей (row_number, name, tag, forms). Вернуть список находок."""
    findings = []
    forms_present = any(f for (_, _, _, f) in rows)  # у «Географ» колонка форм пуста — норма

    # Индексы непустых строк — для детекции пустых строк ВНУТРИ диапазона.
    nonempty_idx = [i for i, (_, n, t, f) in enumerate(rows) if n or t or f]
    last_nonempty = nonempty_idx[-1] if nonempty_idx else -1

    prev_depth = None
    for i, (rn, nm, tag, forms) in enumerate(rows):
        blank = not (nm or tag or forms)

        # 2. blank_row — пустая строка до последней непустой (разрывает стадию [1]).
        if blank:
            if i < last_nonempty:
                findings.append((rn, "blank_row", "пустая строка внутри диапазона данных"))
            continue

        # 3. service_leak — формула/служебная строка в данных.
        if FORMULA.search(nm) or FORMULA.search(forms):
            frag = nm or forms
            findings.append((rn, "service_leak", f"служебная/формульная строка: {frag[:60]}"))
            # не выходим — прочие проверки тоже полезны

        # 7. empty_name — формы есть, термина нет.
        if forms and not nm:
            findings.append((rn, "empty_name", f"формы без термина: {forms[:60]}"))

        # 4. level_gap — скачок уровня >1 относительно предыдущей непустой строки.
        if nm:
            d = depth(nm)
            if prev_depth is not None and d - prev_depth >= 2:
                findings.append(
                    (rn, "level_gap", f"скачок уровня {prev_depth}→{d}: {nm[:50]}")
                )
            prev_depth = d

        # Проверки по колонке форм — только если у листа она вообще заполнена.
        if forms_present and forms:
            # 5. trailing_semicolon
            if forms.endswith(";"):
                findings.append((rn, "trailing_semicolon", f"хвостовой ';': …{forms[-40:]}"))

            segs = split_forms(forms)

            # 1. prose_in_forms — комментарий/проза вместо форм.
            prose_hit = None
            if PROSE_KEYWORDS.search(forms):
                prose_hit = PROSE_KEYWORDS.search(forms).group(0)
            else:
                for s in segs:
                    if len(s.split()) >= LONG_SEGMENT_WORDS or "—" in s:
                        prose_hit = s
                        break
            if prose_hit:
                findings.append(
                    (rn, "prose_in_forms", f"похоже на прозу/комментарий: «{str(prose_hit)[:50]}»")
                )

            # 6. duplicate_forms — повтор формы внутри термина (точный или по регистру).
            seen = {}
            for s in segs:
                k = s.casefold()
                if k in seen:
                    kind = "точный дубль" if seen[k] == s else "дубль по регистру"
                    other = "" if seen[k] == s else f" ≈ «{seen[k][:30]}»"
                    findings.append(
                        (rn, "duplicate_forms", f"{kind}: «{s[:40]}»{other}")
                    )
                    break
                seen[k] = s

    return findings


def load_sheet_rows(ws):
    """Прочитать лист (только первые MAX_COL колонок) → (header, rows)."""
    header = None
    rows = []
    for rn, row in enumerate(
        ws.iter_rows(min_row=1, max_col=MAX_COL, values_only=True), start=1
    ):
        if rn == 1:
            header = row
            continue
        rows.append((rn, row))
    return header, rows


def run(path, only_sheet=None):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    targets = [only_sheet] if only_sheet else WORKING_SHEETS
    results = {}  # sheet -> findings
    for name in targets:
        if name not in wb.sheetnames:
            print(f"⚠ лист «{name}» не найден — пропущен", file=sys.stderr)
            continue
        ws = wb[name]
        header, raw = load_sheet_rows(ws)
        cols = find_columns(header or [])
        if cols["name"] is None and cols["forms"] is None:
            print(f"⚠ лист «{name}»: не опознаны колонки словаря — пропущен", file=sys.stderr)
            continue
        prepared = []
        for rn, row in raw:
            nm = norm(row[cols["name"] - 1]) if cols["name"] else ""
            tg = norm(row[cols["tag"] - 1]) if cols["tag"] else ""
            fm = norm(row[cols["forms"] - 1]) if cols["forms"] else ""
            prepared.append((rn, nm, tg, fm))
        results[name] = check_sheet(name, prepared)
    wb.close()
    return results


def format_report(path, results):
    lines = []
    lines.append(f"# Отчёт валидатора словаря\n")
    lines.append(
        "_Автосгенерировано_ `Litpam-Indexator/tools/validate_dictionary.py` _(H363). "
        "Перегенерировать:_ `python tools/validate_dictionary.py <xlsx> --report <файл>`_._\n"
    )
    lines.append(f"Файл: `{path}`\n")
    total = sum(len(v) for v in results.values())
    # Сводка по классам.
    by_class = {}
    for finds in results.values():
        for _, cls, _ in finds:
            by_class[cls] = by_class.get(cls, 0) + 1
    lines.append("## Сводка\n")
    lines.append(f"Всего находок: **{total}**\n")
    if by_class:
        lines.append("| Класс дефекта | Кол-во |")
        lines.append("|---|---|")
        for cls in sorted(by_class, key=lambda c: -by_class[c]):
            lines.append(f"| `{cls}` | {by_class[cls]} |")
        lines.append("")
    # Подробности.
    for sheet, finds in results.items():
        lines.append(f"## Лист «{sheet}» — {len(finds)} находок\n")
        if not finds:
            lines.append("_чисто_\n")
            continue
        lines.append("| Строка | Класс | Фрагмент |")
        lines.append("|---|---|---|")
        for rn, cls, frag in finds:
            safe = frag.replace("|", "\\|")
            lines.append(f"| {rn} | `{cls}` | {safe} |")
        lines.append("")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(description="Pre-flight-валидатор словаря указателей.")
    ap.add_argument("xlsx", help="путь к .xlsx словаря")
    ap.add_argument("--report", help="записать отчёт в .md файл (иначе — в stdout)")
    ap.add_argument("--sheet", help="проверить только один лист")
    args = ap.parse_args()

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("Ошибка: требуется openpyxl (pip install openpyxl).", file=sys.stderr)
        return 2

    try:
        results = run(args.xlsx, only_sheet=args.sheet)
    except FileNotFoundError:
        print(f"Ошибка: файл не найден: {args.xlsx}", file=sys.stderr)
        return 2

    report = format_report(args.xlsx, results)
    if args.report:
        with open(args.report, "w", encoding="utf-8") as fh:
            fh.write(report + "\n")
        print(f"Отчёт записан: {args.report}", file=sys.stderr)
    else:
        print(report)

    total = sum(len(v) for v in results.values())
    return 1 if total else 0


if __name__ == "__main__":
    sys.exit(main())

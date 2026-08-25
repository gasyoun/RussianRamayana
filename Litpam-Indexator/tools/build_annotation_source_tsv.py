#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_annotation_source_tsv.py -- exports the AddAnnotationData.v.3.jsx
input table (term<TAB>annotation, one pair per line) from the shared
workbook's four index sheets (H2590 remaining stage[4] item: "нужен
двухколоночный источник термин->аннотация" -- confirmed already resolved,
see inspect_annotation_column.py; this script produces the actual TSV).
Only rows with a non-empty "Краткая аннотация" are included; AddAnnotationData
silently skips any term it can't findGrep in the index text, so it is safe
to include every annotated headword from all four sheets rather than
pre-filtering to book II's actual topic set.

Запуск:
    python build_annotation_source_tsv.py [--workbook path.xlsx] --out annotations.tsv

_Автор инструмента: Dr. Mārcis Gasūns · создан 25-08-2026 (H2590 rebuild)._
"""

import argparse
import re
import sys

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

import openpyxl

WORD_CHAR = re.compile(r"\w", re.UNICODE)

DEFAULT_WORKBOOK = "xls/derived/Указатель_к_Рамаяне_1_2_2026_08_15b.xlsx"
SHEETS = ["Именной", "Географ", "Предметы и термины", "Флора и фауна"]


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    pairs = []
    seen_exact = set()
    # AddAnnotationData.v.3.jsx's defaultView findGrep is case-INSENSITIVE
    # ("^(?i)\\bterm\\b"), so two dictionary rows differing only by case (a
    # proper name and an unrelated common-noun homonym, e.g. "Бали" the
    # person vs "бали" the ritual offering -- both real, distinct book-II
    # headwords) both match whichever paragraph the regex finds FIRST,
    # silently concatenating both glosses onto that one topic and leaving
    # the other's paragraph unannotated. Caught live during the H2590
    # rebuild (артефакт: "Бали царь демонов-дайтьев.  жертвенное
    # подношение." on one line). Not something to silently pick a winner
    # for -- both are dropped from the automated table and logged for a
    # human editorial call, same as the =N homonym question the original
    # session parked for MG.
    seen_casefold = {}
    for sheet_name in SHEETS:
        ws = wb[sheet_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        name_idx = header.index("Имя")
        annot_idx = [i for i, h in enumerate(header) if h and "аннотац" in str(h).lower()][0]
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[name_idx] if name_idx < len(row) else None
            annot = row[annot_idx] if annot_idx < len(row) else None
            if not name or not annot:
                continue
            name = str(name).strip()
            annot = str(annot).strip()
            if not name or not annot or name in seen_exact:
                continue
            seen_exact.add(name)
            seen_casefold.setdefault(name.casefold(), []).append((name, annot, sheet_name))
            count += 1
        print(f"{sheet_name}: {count} annotated terms")

    case_collisions = {k: v for k, v in seen_casefold.items() if len(v) > 1}
    for cf, variants in case_collisions.items():
        print(f"  CASE-COLLISION dropped (findGrep is case-insensitive, cannot safely target one): "
              f"{[(n, s) for n, _, s in variants]}")

    survivors = {cf: v[0] for cf, v in seen_casefold.items() if cf not in case_collisions}

    # AddAnnotationData.v.3.jsx's defaultView applies its replacement via
    # changeGrep() with the SAME find pattern used to compute it -- that
    # replaces EVERY document-wide match of "^(?i)\bterm\b", not just the
    # one the operator had in mind. A short headword that is a
    # word-boundary-terminated PREFIX of a longer one (e.g. "Налини" vs
    # "Налини, подруга Анасуи" -- a river and an unrelated person who
    # happens to share the river's name as a prefix, both real book-II
    # topics) matches BOTH paragraphs, so the short term's gloss gets
    # wrongly stamped onto the longer topic too. Caught live during the
    # H2590 rebuild ("Налини река. , подруга Анасуи" -- the malformed
    # concatenation gave it away). Same disposition as a case-collision:
    # drop both, log for a human editorial call.
    names_casefold = sorted(survivors.keys())
    prefix_collisions = set()
    for i, short_cf in enumerate(names_casefold):
        for long_cf in names_casefold:
            if long_cf == short_cf or not long_cf.startswith(short_cf):
                continue
            boundary_char = long_cf[len(short_cf):len(short_cf) + 1]
            if boundary_char and not WORD_CHAR.match(boundary_char):
                prefix_collisions.add(short_cf)
                prefix_collisions.add(long_cf)
                print(f"  PREFIX-COLLISION dropped (changeGrep is document-wide, would double-hit): "
                      f"{survivors[short_cf][0]!r} is a word-boundary prefix of {survivors[long_cf][0]!r}")

    with open(args.out, "w", encoding="utf-8") as f:
        for cf, (name, annot, _sheet) in survivors.items():
            if cf in prefix_collisions:
                continue
            pairs.append((name, annot))
            f.write(f"{name}\t{annot}\n")
    print(f"total: {len(pairs)} pairs -> {args.out} "
          f"({len(case_collisions)} case-collision + {len(prefix_collisions)} prefix-collision headwords excluded)")


if __name__ == "__main__":
    main()

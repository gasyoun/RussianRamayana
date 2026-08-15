"""repair.py — deterministic, reversible correction of validate_dictionary.py findings.

Never edits the source workbook. Writes a new derived copy plus a JSON+MD ledger
recording exactly what changed and why (per-cell old -> new), and what was left
alone as an editorial fork (status WAITING, not silently dropped — H2589 acceptance).
"""

import hashlib
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from validate_dictionary import (  # noqa: E402
    WORKING_SHEETS,
    MAX_COL,
    find_columns,
    load_sheet_rows,
    norm,
    split_forms,
    PROSE_KEYWORDS,
    LONG_SEGMENT_WORDS,
)

FORMS_COLUMN_LABEL = "Что искать, через точку с запятой"


def sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def dedupe_forms(forms):
    """Strip a trailing ';' and drop later case/exact duplicate segments,
    keeping the first-seen casing of each. Returns (new_forms, changed, ops).

    ops is a list of {"kind": "trailing_semicolon"|"duplicate_forms", "detail": str}
    describing what was removed, for the ledger.
    """
    ops = []
    working = forms
    if working.endswith(";"):
        working = working[:-1]
        ops.append({"kind": "trailing_semicolon", "detail": f"хвостовой ';' снят"})

    segs = split_forms(working)
    seen = {}
    kept = []
    for s in segs:
        k = s.casefold()
        if k in seen:
            kind = "точный дубль" if seen[k] == s else "дубль по регистру"
            ops.append(
                {
                    "kind": "duplicate_forms",
                    "detail": f"{kind}: «{s}» удалён (оставлен «{seen[k]}»)",
                }
            )
            continue
        seen[k] = s
        kept.append(s)

    new_forms = "; ".join(kept)
    changed = new_forms != forms
    return new_forms, changed, ops


def is_prose_row(forms):
    """True if the forms cell trips validate_dictionary's prose_in_forms heuristic
    on its FIRST segment specifically (an operator instruction embedded as a form,
    e.g. '[без тега не искать]') — those are an editorial fork, not a dedup target.
    """
    segs = split_forms(forms)
    if not segs:
        return False
    first = segs[0]
    if PROSE_KEYWORDS.search(first):
        return True
    if len(first.split()) >= LONG_SEGMENT_WORDS or first.startswith("["):
        return True
    return False


def compute_operations(source_path, prose_clear=None, ruling_note=None):
    """Read source_path fresh and compute the candidate correction operations.
    Pure/read-only: does not touch the workbook. Returns (source_sha, operations)
    where each operation still carries its 'old' value as read from source_path —
    apply_operations() re-checks that value at apply time (guards staleness if the
    two steps are ever split across time/processes).

    prose_clear: optional set of (sheet_name, row_number) pairs a HUMAN ruled may
    be cleared (the prose operator note is deleted from the forms column). Prose
    rows NOT in the whitelist stay WAITING as before — the ruling is per-cell,
    never a blanket license. ruling_note (who ruled, when) is required with a
    non-empty prose_clear and is recorded verbatim in each cleared op's rationale.
    """
    prose_clear = set(prose_clear or [])
    if prose_clear and not ruling_note:
        raise ValueError("prose_clear requires ruling_note (who ruled, when)")
    source_path = Path(source_path)
    source_sha = sha256_file(source_path)
    wb = openpyxl.load_workbook(source_path, data_only=False)
    operations = []
    seq = 0

    for sheet_name in WORKING_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = [ws.cell(row=1, column=c).value for c in range(1, min(ws.max_column, MAX_COL) + 1)]
        cols = find_columns(header)
        forms_col = cols["forms"]
        if forms_col is None:
            continue

        for row in range(2, ws.max_row + 1):
            raw = norm(ws.cell(row=row, column=forms_col).value)
            if not raw:
                continue

            if is_prose_row(raw):
                seq += 1
                if (sheet_name, row) in prose_clear:
                    operations.append(
                        {
                            "id": f"COR-{seq:04d}",
                            "sheet": sheet_name,
                            "row": row,
                            "column": FORMS_COLUMN_LABEL,
                            "class": "prose_in_forms",
                            "old": raw,
                            "new": "",
                            "disposition": "fixed",
                            "rationale": (
                                "операторская пометка удалена из колонки форм по человеческому "
                                f"рулингу: {ruling_note}"
                            ),
                        }
                    )
                    continue
                operations.append(
                    {
                        "id": f"COR-{seq:04d}",
                        "sheet": sheet_name,
                        "row": row,
                        "column": FORMS_COLUMN_LABEL,
                        "class": "prose_in_forms",
                        "old": raw,
                        "new": raw,
                        "disposition": "WAITING",
                        "rationale": (
                            "первый сегмент — операторская инструкция, а не форма поиска "
                            "(например '[без тега не искать]'); удаление могло бы молчаливо "
                            "отменить редакционное намерение исключить термин из авто-тегирования "
                            "(ruling 23/29) — редакционный форк, не механический дубль."
                        ),
                    }
                )
                continue

            new_forms, changed, ops = dedupe_forms(raw)
            if not changed:
                continue

            seq += 1
            operations.append(
                {
                    "id": f"COR-{seq:04d}",
                    "sheet": sheet_name,
                    "row": row,
                    "column": FORMS_COLUMN_LABEL,
                    "class": "duplicate_forms" if any(o["kind"] == "duplicate_forms" for o in ops) else "trailing_semicolon",
                    "old": raw,
                    "new": new_forms,
                    "disposition": "fixed",
                    "rationale": "; ".join(o["detail"] for o in ops),
                }
            )

    wb.close()
    return source_sha, operations


def apply_operations(source_path, output_path, operations, expected_source_sha=None):
    """Apply a (possibly pre-authored / previously-computed) operations list to a
    FRESH copy of source_path, verifying at apply time that:
      - expected_source_sha (if given) matches the source file actually on disk
        ('source-hash mismatch' guard);
      - each operation's recorded 'old' still matches the cell's current value
        ('stale old' guard) — a mismatch is downgraded to a non-applied,
        explicitly logged conflict rather than silently overwriting drifted data.

    Returns (wrote_output: bool, applied_ops, conflict_ops).
    """
    source_path = Path(source_path)
    output_path = Path(output_path)

    if output_path.resolve() == source_path.resolve():
        raise ValueError("repair-workbook refuses to write in place; --output must differ from --source")

    actual_sha = sha256_file(source_path)
    if expected_source_sha is not None and actual_sha != expected_source_sha:
        raise RuntimeError(
            f"source-hash mismatch: expected {expected_source_sha}, found {actual_sha} — "
            f"operations were computed against a different version of {source_path}; refusing to apply blind"
        )

    wb = openpyxl.load_workbook(source_path, data_only=False)
    applied_ops = []
    conflict_ops = []

    for op in operations:
        ws = wb[op["sheet"]]
        cell = ws.cell(row=op["row"], column=_find_forms_col(ws))
        current = norm(cell.value)
        if current != op["old"]:
            conflict_ops.append({**op, "disposition": "stale_old_conflict", "actual_current": current})
            continue
        if op["disposition"] == "fixed":
            cell.value = op["new"] if op["new"] != "" else None
        applied_ops.append(op)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    resha = sha256_file(source_path)
    if resha != actual_sha:
        raise RuntimeError("source workbook SHA changed during repair — this must never happen")

    return True, applied_ops, conflict_ops


def _find_forms_col(ws):
    header = [ws.cell(row=1, column=c).value for c in range(1, min(ws.max_column, MAX_COL) + 1)]
    return find_columns(header)["forms"]


def repair_workbook(source_path, output_path, ledger_path, prose_clear=None, ruling_note=None):
    source_sha, operations = compute_operations(source_path, prose_clear=prose_clear, ruling_note=ruling_note)
    _, applied_ops, conflict_ops = apply_operations(source_path, output_path, operations, expected_source_sha=source_sha)

    output_sha = sha256_file(output_path)
    ledger = {
        "handoff": "H2589",
        "tool": "print_ready.py repair-workbook",
        "source_path": str(Path(source_path)).replace("\\", "/"),
        "source_sha256": source_sha,
        "output_path": str(Path(output_path)).replace("\\", "/"),
        "output_sha256": output_sha,
        "operation_count": len(applied_ops) + len(conflict_ops),
        "fixed_count": sum(1 for o in applied_ops if o["disposition"] == "fixed"),
        "waiting_count": sum(1 for o in applied_ops if o["disposition"] == "WAITING"),
        "conflict_count": len(conflict_ops),
        "operations": applied_ops + conflict_ops,
    }
    ledger_path = Path(ledger_path)
    ledger_path.parent.mkdir(parents=True, exist_ok=True)
    ledger_path.write_text(json.dumps(ledger, ensure_ascii=False, indent=2), encoding="utf-8")

    return ledger


def format_ledger_md(ledger):
    lines = []
    lines.append("# Correction ledger — Litpam-Indexator dictionary repair (H2589)\n")
    lines.append(f"Source: `{ledger['source_path']}` (SHA-256 `{ledger['source_sha256']}`, unchanged)\n")
    lines.append(f"Output: `{ledger['output_path']}` (SHA-256 `{ledger['output_sha256']}`)\n")
    lines.append(f"Operations: {ledger['operation_count']} (`fixed`: {ledger['fixed_count']}, `WAITING`: {ledger['waiting_count']})\n")
    lines.append("| ID | Sheet | Row | Class | Disposition | Old | New | Rationale |")
    lines.append("|---|---|---|---|---|---|---|---|")
    for op in ledger["operations"]:
        old = op["old"].replace("|", "\\|")[:80]
        new = op["new"].replace("|", "\\|")[:80]
        rationale = op["rationale"].replace("|", "\\|")
        lines.append(
            f"| {op['id']} | {op['sheet']} | {op['row']} | {op['class']} | {op['disposition']} | {old} | {new} | {rationale} |"
        )
    return "\n".join(lines) + "\n"

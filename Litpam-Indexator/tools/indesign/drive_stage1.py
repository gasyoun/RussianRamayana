#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
drive_stage1.py — COM driver for MANUAL stage [1] (H2776, MG override 15-08-2026:
«Попробуй стадию [1] сам через COM»).

For each of the four workbook sheets (canonical order fixes the IndexList numbers:
Именной→001, Географ→002, Предметы и термины→003, Флора и фауна→004):

  1. extract columns «Имя» + «Что искать» to a TSV (blank rows dropped);
  2. build_indexlist_table.jsx creates the 2-column table document with the
     #IndexStyles group, threaded frames (no overset), cursor in first cell;
  3. run the AUTHORIAL UseReadyTable.v.7.jsx unmodified via DoScript — its modal
     alert()s are redirected to $.global.__alertLog inside its own persistent
     "UseReadyTable" engine beforehand, so nothing blocks headless;
  4. verify the autonumbered IndexList-NNN.indd appeared, clear residual overset,
     close docs, log per-sheet results.

Everything runs inside work/print-readiness/ (outside Git); authorial scripts are
never edited (ruling 26 / IMPLEMENTATION step-6 item 3 as overridden by MG for
stage [1] wrappers).

Запуск:
    python drive_stage1.py [--sheets "Именной,Географ,..."] [--workdir <dir>] [--quit]

_Автор инструмента: Dr. Mārcis Gasūns · создан 15-08-2026 (H2776)._
"""

import argparse
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

import openpyxl  # noqa: E402

ID_JAVASCRIPT = 1246973031
ID_NEVER_INTERACT = 1699640946
SAVE_OPTIONS_NO = 1852776480

HERE = Path(__file__).resolve().parent
BASE = HERE.parents[1]  # Litpam-Indexator/
BUILD_JSX = HERE / "build_indexlist_table.jsx"
USE_READY = BASE / "#Indexing. Ramayana" / "[1. Подготовка таблиц]" / "UseReadyTable.v.7.jsx"
WORKBOOK = BASE / "work" / "print-readiness" / "pilot-I" / "Указатель_к_Рамаяне_1_2_2026_08_15.xlsx"

SHEETS = ["Именной", "Географ", "Предметы и термины", "Флора и фауна"]

ALERT_SHIM = (
    '#targetengine "UseReadyTable"\r'
    "alert = function (m) { $.global.__alertLog = ($.global.__alertLog || \"\") + m + \" || \"; };\r"
    '"shim-ok"'
)
ALERT_READ = '#targetengine "UseReadyTable"\rvar r = $.global.__alertLog || ""; $.global.__alertLog = ""; r'


def find_columns(header):
    name_col = forms_col = None
    for i, h in enumerate(header, start=1):
        hl = str(h or "").strip().lower()
        if name_col is None and (hl.startswith("имя") or "теги из текста" in hl):
            name_col = i
        elif forms_col is None and "что искать" in hl:
            forms_col = i
    return name_col, forms_col


def sheet_to_tsv(wb, sheet, out_path):
    ws = wb[sheet]
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    name_col, forms_col = find_columns(header)
    if not name_col or not forms_col:
        raise SystemExit(f"{sheet}: cannot locate Имя / Что искать columns in header {header[:6]}")
    lines = []
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(row=r, column=name_col).value or "").strip()
        forms = str(ws.cell(row=r, column=forms_col).value or "").strip()
        if not name and not forms:
            continue
        name = " ".join(name.split("\n"))
        forms = " ".join(forms.split("\n"))
        lines.append(f"{name}\t{forms}")
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return len(lines)


def main(argv=None):
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--sheets", default=",".join(SHEETS))
    p.add_argument("--workdir", default=str(BASE / "work" / "print-readiness" / "pilot-I" / "indexlists"))
    p.add_argument("--quit", action="store_true")
    args = p.parse_args(argv)

    workdir = Path(args.workdir).resolve()
    if "work" not in workdir.parts or "print-readiness" not in workdir.parts:
        raise SystemExit("refusing: --workdir must live under work/print-readiness/ (ruling 26)")
    workdir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)

    import win32com.client

    app = win32com.client.Dispatch("InDesign.Application")
    print(f"[stage1] connected: {app.Name} {app.Version}")
    app.ScriptPreferences.UserInteractionLevel = ID_NEVER_INTERACT
    shim = app.DoScript(ALERT_SHIM, ID_JAVASCRIPT)
    print(f"[stage1] alert shim installed in engine UseReadyTable: {shim}")

    results = []
    for sheet in [s.strip() for s in args.sheets.split(",") if s.strip()]:
        slug = {"Именной": "imennoj", "Географ": "geograf", "Предметы и термины": "predmety", "Флора и фауна": "flora"}.get(sheet, "sheet")
        tsv = workdir / f"src-{slug}.tsv"
        n = sheet_to_tsv(wb, sheet, tsv)
        print(f"[stage1] {sheet}: {n} data rows -> {tsv.name}")

        before = {f.name for f in workdir.glob("IndexList*.indd")}

        src_indd = workdir / f"Source-{slug}.indd"
        if src_indd.exists():
            src_indd.unlink()
        app.ScriptArgs.SetValue("tsvPath", str(tsv))
        app.ScriptArgs.SetValue("saveAsPath", str(src_indd))
        t0 = time.time()
        build_out = app.DoScript(str(BUILD_JSX), ID_JAVASCRIPT)
        print(f"[stage1] {sheet}: table built in {time.time()-t0:.0f}s — {build_out}")

        t0 = time.time()
        try:
            app.DoScript(str(USE_READY), ID_JAVASCRIPT)
        except Exception as e:  # noqa: BLE001 — the authorial script may abort via its own error paths
            print(f"[stage1] {sheet}: UseReadyTable raised: {e}")
        alerts = app.DoScript(ALERT_READ, ID_JAVASCRIPT)
        after = {f.name for f in workdir.glob("IndexList*.indd")}
        new_files = sorted(after - before)
        print(
            f"[stage1] {sheet}: UseReadyTable finished in {time.time()-t0:.0f}s; "
            f"new file(s): {new_files or 'NONE'}; alerts: {alerts or '(none)'}"
        )
        results.append((sheet, n, new_files, str(alerts)))

        # Close every doc still open (the working doc was re-saved as IndexList-NNN
        # by the authorial script on success), never saving on close.
        while app.Documents.Count > 0:
            app.Documents.Item(1).Close(SAVE_OPTIONS_NO)

    if args.quit:
        app.Quit(SAVE_OPTIONS_NO)

    print("\n[stage1] summary:")
    ok = True
    for sheet, n, new_files, alerts in results:
        status = "OK" if new_files and "Таблица сохранена" in alerts else "CHECK"
        if status != "OK":
            ok = False
        print(f"  {status}: {sheet} ({n} rows) -> {', '.join(new_files) or '—'} | {alerts[:200]}")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())

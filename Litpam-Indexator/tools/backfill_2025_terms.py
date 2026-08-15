#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
backfill_2025_terms.py — внести в словник термины указателей 2025 года, которых
в нём не было (рулинг МГ 15-08-2026: «Backfill вариант (а) — вноси в .xlsx»).

Куратура: из 29 механических кандидатов (build_backfill_list.py) вручную отсеяны
обломки переносов; регистр и принадлежность к указателю восстановлены по живым
строкам страниц 415–438 пруфа 2025 (постраничная атрибуция). Примечательное:
Вриттимана/Дхану/Дханью/Дхритималина — имена ОРУЖИЯ из даров Вишвамитры
(стр. 102 текста), потому стоят в предметном указателе; «ашрама» была в ДВУХ
указателях (географическом p430 и предметном p431) — вносится в оба листа.

Падежные формы — pymorphy3 через существующий gen_case_forms.term_forms (H377).
Исходный дериват НЕ мутируется: пишется новая версия + ledger.

Запуск:
    python backfill_2025_terms.py --source <derived.xlsx> --output <new.xlsx> --ledger <ledger.json>

_Автор инструмента: Dr. Mārcis Gasūns · создан 15-08-2026 (H2776)._
"""

import argparse
import hashlib
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

import openpyxl  # noqa: E402

from gen_case_forms import get_analyzer, term_forms  # noqa: E402

# (лист, термин-как-в-2025, страница-2025-указателя, примечание)
CURATED = [
    ("Именной", "царь (владыка) Кекаи", 416, "в 2025 — статья именного; вероятный подуровень Ашвапати, оператор уточнит"),
    ("Именной", "Пашупати-рудра", 426, ""),
    ("Географ", "Рикшавата", 430, "гора"),
    ("Географ", "ашрама", 430, "в 2025 присутствовала и в географическом"),
    ("Предметы и термины", "адитьи", 431, "класс существ"),
    ("Предметы и термины", "асуры", 431, "класс существ"),
    ("Предметы и термины", "ашрама", 431, ""),
    ("Предметы и термины", "данавы", 432, "класс существ"),
    ("Предметы и термины", "Вриттимана", 433, "имя оружия (дары Вишвамитры)"),
    ("Предметы и термины", "Дхану", 433, "имя оружия"),
    ("Предметы и термины", "Дханью", 433, "имя оружия"),
    ("Предметы и термины", "Дхритималина", 433, "имя оружия"),
    ("Предметы и термины", "небо Индры", 433, ""),
    ("Предметы и термины", "лук Шивы", 434, ""),
    ("Предметы и термины", "Праджапати Кришашвы", 434, ""),
    ("Предметы и термины", "пурана", 435, ""),
    ("Предметы и термины", "Уттара Пхалгуни", 438, "накшатра; p438 — продолжение предметного"),
    ("Предметы и термины", "Южный Крест", 438, "созвездие; p438 — продолжение предметного"),
]

RULING = "рулинг МГ 15-08-2026: «Backfill вариант (а) — вноси в .xlsx и гони стадию [3]»"


def sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def last_data_row(ws):
    for r in range(ws.max_row, 1, -1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, min(ws.max_column, 12) + 1)):
            return r
    return 1


def find_forms_col(ws):
    for c in range(1, min(ws.max_column, 30) + 1):
        h = str(ws.cell(row=1, column=c).value or "").lower()
        if "что искать" in h and "заготовка" not in h:
            return c
    raise SystemExit(f"{ws.title}: no «Что искать» column")


def main(argv=None):
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--source", required=True)
    p.add_argument("--output", required=True)
    p.add_argument("--ledger", required=True)
    args = p.parse_args(argv)

    source = Path(args.source)
    output = Path(args.output)
    if output.resolve() == source.resolve():
        raise SystemExit("refusing to write in place")
    source_sha = sha256_file(source)

    morph = get_analyzer()
    wb = openpyxl.load_workbook(source)
    entries = []
    for sheet, term, page, note in CURATED:
        ws = wb[sheet]
        forms_col = find_forms_col(ws)
        forms = term_forms(morph, term)
        # дедуп с сохранением порядка; сам термин — первой формой
        uniq = []
        for f in [term] + forms:
            if f not in uniq:
                uniq.append(f)
        forms_str = "; ".join(uniq)
        row = last_data_row(ws) + 1
        ws.cell(row=row, column=1, value=term)
        ws.cell(row=row, column=forms_col, value=forms_str)
        entries.append(
            {
                "sheet": sheet, "row": row, "term": term, "forms": forms_str,
                "source_2025_page": page, "note": note,
            }
        )
        print(f"[backfill] {sheet} row {row}: {term} ({len(uniq)} форм)")

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    ledger = {
        "tool": "backfill_2025_terms.py (H2776)",
        "ruling": RULING,
        "source_path": str(source).replace("\\", "/"),
        "source_sha256": source_sha,
        "output_path": str(output).replace("\\", "/"),
        "output_sha256": sha256_file(output),
        "entries": entries,
        "curation_note": (
            "18 строк из 29 механических кандидатов; отсев обломков переносов и "
            "form-differences задокументирован в backfill-candidates.json (excluded)."
        ),
    }
    Path(args.ledger).parent.mkdir(parents=True, exist_ok=True)
    Path(args.ledger).write_text(json.dumps(ledger, ensure_ascii=False, indent=1), encoding="utf-8")
    resha = sha256_file(source)
    assert resha == source_sha, "source workbook changed during backfill — must never happen"
    print(f"[backfill] {len(entries)} rows -> {output}\nledger: {args.ledger}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

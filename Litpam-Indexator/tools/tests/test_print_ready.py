"""Tests for print_ready.py and the print_ready package (H2589).

Covers: dedupe/repair correctness + idempotence, source-hash mismatch, stale-old
conflict, missing-status/prose-fork handling, audit-idml against a synthetic IDML,
audit-pdf against a synthetic PDF, verify-packet PASS/FAIL (missing member, broken
checksum), and coverage's headword extraction.
"""

import io
import json
import zipfile

import openpyxl
import pytest

from print_ready import repair as repair_mod
from print_ready import idml_audit
from print_ready import pdf_audit
from print_ready import packet as packet_mod
from print_ready import coverage as coverage_mod


# --- repair.dedupe_forms -----------------------------------------------------

def test_dedupe_forms_removes_case_duplicate_keeps_first():
    new, changed, ops = repair_mod.dedupe_forms("Идущий впереди; идущим впереди; идущий впереди")
    assert changed
    assert new == "Идущий впереди; идущим впереди"
    assert any(o["kind"] == "duplicate_forms" for o in ops)


def test_dedupe_forms_removes_exact_duplicate():
    new, changed, ops = repair_mod.dedupe_forms("сын Вивасвата; сын Вивасвата; сына Вивасвата")
    assert changed
    assert new == "сын Вивасвата; сына Вивасвата"


def test_dedupe_forms_strips_trailing_semicolon():
    new, changed, ops = repair_mod.dedupe_forms("Пашупати;")
    assert changed
    assert new == "Пашупати"
    assert any(o["kind"] == "trailing_semicolon" for o in ops)


def test_dedupe_forms_noop_when_clean():
    new, changed, ops = repair_mod.dedupe_forms("апсара; апсары; апсарам")
    assert not changed
    assert new == "апсара; апсары; апсарам"
    assert ops == []


def test_is_prose_row_flags_bracketed_instruction():
    assert repair_mod.is_prose_row("[без тега не искать]; оружие; оружием")


def test_is_prose_row_false_for_ordinary_forms():
    assert not repair_mod.is_prose_row("владыка Кекаи; владыке Кекаи")


# --- repair_workbook end-to-end on a tiny synthetic workbook -----------------

def _make_workbook(rows, sheet="Именной"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["Имя", "Тег для поиска в тексте", "Что искать, через точку с запятой"])
    for name, forms in rows:
        ws.append([name, "", forms])
    return wb


def test_repair_workbook_end_to_end_and_idempotent(tmp_path):
    wb = _make_workbook(
        [
            ("Термин1", "Форма; форма; форма2"),
            ("Термин2", "чисто; уже без дублей"),
            ("Оружие", "[без тега не искать]; оружие; оружием"),
        ]
    )
    source = tmp_path / "source.xlsx"
    wb.save(source)
    source_bytes_before = source.read_bytes()

    output1 = tmp_path / "derived1.xlsx"
    ledger1 = tmp_path / "ledger1.json"
    result = repair_mod.repair_workbook(source, output1, ledger1)

    assert result["fixed_count"] == 1  # Термин1 loses one case-dup
    assert result["waiting_count"] == 1  # Оружие stays WAITING
    assert result["conflict_count"] == 0
    assert source.read_bytes() == source_bytes_before  # never touch the source

    wb1 = openpyxl.load_workbook(output1)
    ws1 = wb1["Именной"]
    assert ws1.cell(row=2, column=3).value == "Форма; форма2"
    assert ws1.cell(row=4, column=3).value == "[без тега не искать]; оружие; оружием"  # untouched

    # Idempotence: repairing the already-repaired copy makes zero further fixes.
    output2 = tmp_path / "derived2.xlsx"
    ledger2 = tmp_path / "ledger2.json"
    result2 = repair_mod.repair_workbook(output1, output2, ledger2)
    assert result2["fixed_count"] == 0
    assert result2["waiting_count"] == 1  # the editorial fork is still reported, not silently resolved


def test_apply_operations_source_hash_mismatch(tmp_path):
    wb = _make_workbook([("Термин1", "форма; форма")])
    source = tmp_path / "source.xlsx"
    wb.save(source)
    _, ops = repair_mod.compute_operations(source)

    with pytest.raises(RuntimeError, match="source-hash mismatch"):
        repair_mod.apply_operations(source, tmp_path / "out.xlsx", ops, expected_source_sha="deadbeef" * 8)


def test_apply_operations_stale_old_is_logged_not_silently_applied(tmp_path):
    wb = _make_workbook([("Термин1", "форма; форма")])
    source = tmp_path / "source.xlsx"
    wb.save(source)
    _, ops = repair_mod.compute_operations(source)

    # Drift the source after operations were computed against it.
    wb2 = openpyxl.load_workbook(source)
    wb2["Именной"].cell(row=2, column=3).value = "форма; другая форма"
    wb2.save(source)

    _, applied, conflicts = repair_mod.apply_operations(source, tmp_path / "out.xlsx", ops)
    assert len(conflicts) == 1
    assert conflicts[0]["disposition"] == "stale_old_conflict"
    assert applied == []  # nothing silently overwritten on drifted data


def test_repair_workbook_refuses_in_place(tmp_path):
    wb = _make_workbook([("T", "a; a")])
    source = tmp_path / "source.xlsx"
    wb.save(source)
    with pytest.raises(ValueError, match="in place"):
        repair_mod.repair_workbook(source, source, tmp_path / "ledger.json")


# --- audit-idml ----------------------------------------------------------------

def _make_synthetic_idml(path):
    designmap = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<?aid style="50" type="document" readerVersion="6.0" featureSet="257" product="17.4(51)" ?>\n'
        '<Document xmlns:idPkg="http://ns.adobe.com/AdobeInDesign/idml/1.0/packaging" '
        'DOMVersion="17.0" Self="d" StoryList="u1 u2">\n'
        '  <idPkg:Fonts src="Resources/Fonts.xml" />\n'
        '  <idPkg:Spread src="Spreads/Spread_u1.xml" />\n'
        '  <idPkg:Story src="Stories/Story_u1.xml" />\n'
        '  <Section Self="sec1" Name="" Marker="" PageNumberStart="1" PageNumberType="Arabic" />\n'
        "</Document>\n"
    )
    spread = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<idPkg:Spread xmlns:idPkg="http://ns.adobe.com/AdobeInDesign/idml/1.0/packaging" DOMVersion="17.0">\n'
        '  <Spread Self="u1" PageCount="1">\n'
        '    <Page Self="p1" Name="1" GeometricBounds="0 0 609.4488188976001 467.71653543310003" AppliedMaster="m1" />\n'
        '  </Spread>\n'
        "</idPkg:Spread>\n"
    )
    story = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<idPkg:Story xmlns:idPkg="http://ns.adobe.com/AdobeInDesign/idml/1.0/packaging" DOMVersion="17.0">\n'
        '  <Story Self="u1">\n'
        '    <ParagraphStyleRange>\n'
        '      <CharacterStyleRange><Content>Индра — владыка богов</Content></CharacterStyleRange>\n'
        "    </ParagraphStyleRange>\n"
        "  </Story>\n"
        "</idPkg:Story>\n"
    )
    fonts = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<idPkg:Fonts xmlns:idPkg="http://ns.adobe.com/AdobeInDesign/idml/1.0/packaging" DOMVersion="17.0">\n'
        '  <FontFamily Self="ff1" Name="NewBaskervilleITC">\n'
        '    <Font Self="f1" Name="NewBaskervilleITC-Roman" PostScriptName="NewBaskervilleITC-Roman" '
        'FontType="OpenTypeCFF" Status="Installed" />\n'
        "  </FontFamily>\n"
        "</idPkg:Fonts>\n"
    )
    with zipfile.ZipFile(path, "w") as z:
        z.writestr("mimetype", "application/vnd.adobe.indesign-idml-package")
        z.writestr("designmap.xml", designmap)
        z.writestr("Spreads/Spread_u1.xml", spread)
        z.writestr("Stories/Story_u1.xml", story)
        z.writestr("Resources/Fonts.xml", fonts)


def test_audit_idml_structural_facts(tmp_path):
    idml_path = tmp_path / "sample.idml"
    _make_synthetic_idml(idml_path)

    result = idml_audit.audit(idml_path)
    assert result["product_version"] == "17.4(51)"
    assert result["dom_version"] == "17.0"
    assert result["page_count"] == 1
    assert result["pages"][0]["geometric_bounds"] == "0 0 609.4488188976001 467.71653543310003"
    assert len(result["fonts"]) == 1
    assert result["fonts"][0]["status"] == "Installed"
    assert len(result["stories"]) == 1
    assert "Индра" in "" or result["stories"][0]["text_length"] > 0
    assert len(result["sections"]) == 1


# --- audit-pdf -------------------------------------------------------------

def test_audit_pdf_page_facts(tmp_path):
    fitz = pytest.importorskip("fitz")
    pdf_path = tmp_path / "sample.pdf"
    doc = fitz.open()
    page = doc.new_page(width=467.72, height=609.45)
    page.insert_text((50, 50), "УКАЗАТЕЛЬ ИМЕН")
    doc.save(pdf_path)
    doc.close()

    result = pdf_audit.audit(pdf_path)
    assert result["page_count"] == 1
    assert result["pages"][0]["text_length"] > 0
    assert result["dependency"]["name"] == "PyMuPDF"


# --- verify-packet -----------------------------------------------------------

def test_verify_packet_pass(tmp_path):
    member = tmp_path / "evidence.json"
    member.write_text('{"a": 1}', encoding="utf-8")
    sha = repair_mod.sha256_file(member)
    manifest = {
        "packet_id": "book-I-baseline-2022",
        "handoff": "H2589",
        "created": "2026-08-14",
        "members": [{"path": "evidence.json", "sha256": sha}],
    }
    (tmp_path / "manifest.json").write_text(json.dumps(manifest), encoding="utf-8")

    ok, report = packet_mod.verify(tmp_path)
    assert ok
    assert "PASS: packet verified" in report


def test_verify_packet_fails_on_missing_member(tmp_path):
    manifest = {
        "packet_id": "book-I-baseline-2022",
        "handoff": "H2589",
        "created": "2026-08-14",
        "members": [{"path": "missing.json", "sha256": "0" * 64}],
    }
    (tmp_path / "manifest.json").write_text(json.dumps(manifest), encoding="utf-8")

    ok, report = packet_mod.verify(tmp_path)
    assert not ok
    assert "missing packet member" in report


def test_verify_packet_fails_on_checksum_mismatch(tmp_path):
    member = tmp_path / "evidence.json"
    member.write_text('{"a": 1}', encoding="utf-8")
    manifest = {
        "packet_id": "book-I-baseline-2022",
        "handoff": "H2589",
        "created": "2026-08-14",
        "members": [{"path": "evidence.json", "sha256": "0" * 64}],
    }
    (tmp_path / "manifest.json").write_text(json.dumps(manifest), encoding="utf-8")

    ok, report = packet_mod.verify(tmp_path)
    assert not ok
    assert "checksum mismatch" in report


def test_verify_packet_fails_on_missing_manifest(tmp_path):
    ok, report = packet_mod.verify(tmp_path)
    assert not ok
    assert "manifest.json missing" in report


# --- coverage.headword_text --------------------------------------------------

@pytest.mark.parametrize(
    "name,expected",
    [
        ("Агни\\идущий впереди [богов]", "идущий впереди [богов]"),
        ("Брахма\\Сваямбху, или Самосущий", "Сваямбху"),
        ("Корова, исполняющая желания", "Корова, исполняющая желания"),
        ("Индра\\Пурандара, или Сокрушитель крепостей, или Покоритель городов", "Пурандара"),
    ],
)
def test_headword_text(name, expected):
    assert coverage_mod.headword_text(name) == expected

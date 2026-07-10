#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
enrich_dictionary.py — DeepSeek-обогащение словника указателей «Рамаяны» (H377, роль a).

Поверх строго read-only `../validate_dictionary.py`: для каждого термина рабочего
листа словаря (`xls/Указатель_к_Рамаяне_*.xlsx`) LLM ПРОВЕРЯЕТ колонку
«Что искать в тексте» (падежные формы через «;») и ПРЕДЛАГАЕТ недостающие формы —
замену ручному склонятелю `getWord()`/`procFamily()` из `ForIndex.jsxinc`.

Философия та же, что у валидатора: сам .xlsx НЕ меняется. Результат — отчёт
`dictionary-enrichment-report.md` рядом со словником (или --report), в котором
по каждому термину: вердикт ok / missing_forms / suspicious + предлагаемый
полный набор форм. Оператор переносит правки в xlsx руками.

Бюджет: $20 по умолчанию (--budget), running-стоимость печатается после каждого
вызова, останов на 95% лимита (deepseek_common.BudgetExceeded) — уже обработанные
термины при этом сохраняются в отчёт.

Запуск:
    python enrich_dictionary.py <путь-к-xlsx>
    python enrich_dictionary.py <xlsx> --sheet "Именной" --limit 50
    python enrich_dictionary.py <xlsx> --budget 5 --report отчёт.md

Коды возврата: 0 — все термины ok; 1 — есть предложения/подозрения; 2 — ошибка.

Зависимости: openpyxl (как у validate_dictionary.py) + stdlib.

_Автор: Dr. Mārcis Gasūns · создан 10-07-2026 (H377)._
"""

import argparse
import datetime
import json
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from deepseek_common import BudgetExceeded, DeepSeekClient  # noqa: E402

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
from validate_dictionary import WORKING_SHEETS, find_columns, norm  # noqa: E402

SYSTEM_PROMPT = (
    "Ты — лингвист-русист, помогаешь составителю предметного указателя к русскому "
    "переводу «Рамаяны». Для термина указателя дан список словоформ, по которым "
    "термин ищется в тексте (падежные формы, варианты). Проверь список:\n"
    "1) все ли реально употребимые падежные формы (ед. и мн. числа, если уместно) "
    "покрыты;\n"
    "2) нет ли в списке ошибочных/несуществующих форм;\n"
    "3) для несклоняемых слов список из одной формы — норма.\n"
    "Имена собственные санскритского происхождения (Рама, Сита, Хануман, Айодхья…) "
    "склоняй по правилам русского языка, как они склоняются в академических "
    "переводах. Отвечай СТРОГО одним JSON-объектом:\n"
    '{"verdict": "ok" | "missing_forms" | "suspicious", '
    '"forms": "полный предлагаемый список форм через ; (базовые формы, без '
    'регекс-синтаксиса)", "note": "одна строка пояснения по-русски"}'
)


def iter_terms(ws, max_rows=None):
    """(row_idx, term, forms) по строкам листа; колонки ищутся по заголовку."""
    header = [c.value for c in ws[1]]
    cols = find_columns(header)
    if cols["name"] is None or cols["forms"] is None:
        return
    count = 0
    for r in range(2, ws.max_row + 1):
        term = norm(ws.cell(row=r, column=cols["name"]).value)
        forms = norm(ws.cell(row=r, column=cols["forms"]).value)
        if not term:
            continue
        yield r, term, forms
        count += 1
        if max_rows and count >= max_rows:
            return


def ask(client, term, forms):
    user = "Термин: %s\nТекущие формы: %s" % (term, forms or "(пусто)")
    raw = client.chat(SYSTEM_PROMPT, user, max_tokens=400)
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.strip("`")
        raw = raw[raw.find("{"):raw.rfind("}") + 1]
    try:
        return json.loads(raw)
    except ValueError:
        return {"verdict": "suspicious", "forms": forms,
                "note": "ответ LLM не разобран: " + raw[:120]}


def main(argv=None):
    ap = argparse.ArgumentParser(description="DeepSeek-обогащение словника указателей")
    ap.add_argument("xlsx", help="путь к Указатель_к_Рамаяне_*.xlsx")
    ap.add_argument("--sheet", help="только один рабочий лист")
    ap.add_argument("--limit", type=int, default=0, help="максимум терминов (0 = все)")
    ap.add_argument("--budget", type=float, default=None, help="бюджет прогона, $")
    ap.add_argument("--report", help="путь к md-отчёту (по умолчанию рядом с xlsx)")
    args = ap.parse_args(argv)

    try:
        import openpyxl
    except ImportError:
        print("Нужен openpyxl: pip install openpyxl", file=sys.stderr)
        return 2

    if not os.path.isfile(args.xlsx):
        print("Файл не найден: %s" % args.xlsx, file=sys.stderr)
        return 2

    client = DeepSeekClient(budget_usd=args.budget)
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    sheets = [args.sheet] if args.sheet else [s for s in WORKING_SHEETS if s in wb.sheetnames]

    rows_out = []
    n_flagged = 0
    stopped = ""
    try:
        for sheet in sheets:
            if sheet not in wb.sheetnames:
                print("Листа нет: %s" % sheet, file=sys.stderr)
                continue
            for r, term, forms in iter_terms(wb[sheet], args.limit or None):
                verdict = ask(client, term, forms)
                if verdict.get("verdict") != "ok":
                    n_flagged += 1
                rows_out.append((sheet, r, term, forms, verdict))
    except BudgetExceeded as e:
        stopped = str(e)
        print(stopped, file=sys.stderr)

    report = args.report or os.path.join(
        os.path.dirname(os.path.abspath(args.xlsx)), "dictionary-enrichment-report.md")
    today = datetime.date.today().strftime("%d-%m-%Y")
    with open(report, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("# Отчёт DeepSeek-обогащения словника\n\n")
        fh.write("_Created: %s · Last updated: %s_\n\n" % (today, today))
        fh.write("Источник: `%s` · модель: `%s` · %s\n\n"
                 % (os.path.basename(args.xlsx), client.model, client.cost_line()))
        if stopped:
            fh.write("**Прогон остановлен бюджетным стопом:** %s\n\n" % stopped)
        fh.write("| Лист | Строка | Термин | Вердикт | Предлагаемые формы | Пояснение |\n")
        fh.write("|---|---|---|---|---|---|\n")
        for sheet, r, term, forms, v in rows_out:
            fh.write("| %s | %d | %s | %s | %s | %s |\n" % (
                sheet, r, term, v.get("verdict", "?"),
                v.get("forms", "").replace("|", "/"),
                v.get("note", "").replace("|", "/")))
        fh.write("\n_Сгенерировано `tools/copilot/enrich_dictionary.py`; xlsx не"
                 " изменялся — правки переносит оператор._\n")
    print("Отчёт: %s · терминов: %d · с предложениями: %d"
          % (report, len(rows_out), n_flagged))
    print(client.cost_line())
    return 1 if (n_flagged or stopped) else 0


if __name__ == "__main__":
    sys.exit(main())

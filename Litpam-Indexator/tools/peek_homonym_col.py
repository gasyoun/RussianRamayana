#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Peek 'Для Омонимов' column contents for rows matching known print subentry
cases (Аншуман, Бали, Агни/Анала/Павака), to see if it's the source data for
AddAnnotationData subentry glosses."""
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

wb = openpyxl.load_workbook(
    "xls/derived/Указатель_к_Рамаяне_1_2_2026_08_15b.xlsx", data_only=True
)
ws = wb["Именной"]
header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
print("header:", [h for h in header if h])
col_map = {str(h).strip(): i for i, h in enumerate(header, start=1) if h}

TARGETS = {"Аншуман", "Бали", "Агни", "Анала", "Павака", "Аламбуша", "Агастья"}
for r in range(2, ws.max_row + 1):
    name = str(ws.cell(row=r, column=col_map["Имя"]).value or "").strip()
    if name in TARGETS:
        row_vals = {h: ws.cell(row=r, column=c).value for h, c in col_map.items()}
        print(f"row {r}: {name!r}")
        for h in ["Что искать, через точку с запятой", "Описание", "Краткая аннотация", "Для Омонимов", "Источник описания"]:
            if h in row_vals:
                print(f"    {h}: {row_vals[h]!r}")

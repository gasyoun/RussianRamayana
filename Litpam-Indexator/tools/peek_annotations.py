#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Quick peek: count non-empty 'Краткая аннотация' cells per sheet, print samples."""
import sys
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

wb = openpyxl.load_workbook(
    "xls/derived/Указатель_к_Рамаяне_1_2_2026_08_15b.xlsx", data_only=True
)
SHEETS = ["Именной", "Географ", "Предметы и термины", "Флора и фауна"]
for sheet in SHEETS:
    ws = wb[sheet]
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    name_col = None
    ann_col = None
    for i, h in enumerate(header, start=1):
        hl = str(h or "").strip().lower()
        if hl.startswith("имя"):
            name_col = i
        if hl == "краткая аннотация":
            ann_col = i
    print(f"=== {sheet}: name_col={name_col} ann_col={ann_col} ===")
    if not ann_col:
        continue
    n = 0
    samples = []
    for r in range(2, ws.max_row + 1):
        ann = ws.cell(row=r, column=ann_col).value
        if ann and str(ann).strip():
            n += 1
            if len(samples) < 10:
                name = ws.cell(row=r, column=name_col).value
                samples.append((r, name, ann))
    print(f"  non-empty annotations: {n}")
    for r, name, ann in samples:
        print(f"    row {r}: {name!r} -> {ann!r}")

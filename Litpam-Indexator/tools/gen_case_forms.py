#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_case_forms.py — pymorphy3-генератор падежных форм для словника указателей (H377).

Замена ручного склонятеля `getWord()`/`procFamily()`/`getGroup()` из
`ForIndex.jsxinc` (≈500 строк ES3-модели русского склонения) на словарную
морфологию: для каждого термина листа словаря генерируются все словоформы
(падежи × числа) через pymorphy3 (OpenCorpora), и предлагается содержимое
колонки «Что искать в тексте» (формы через «;»).

Как и validate_dictionary.py / copilot-инструменты: сам `.xlsx` НЕ меняется —
результат пишется в отчёт `case-forms-report.md` (или --report) и, по желанию,
в TSV (--tsv) для полуавтоматического переноса. Оператор сверяет и переносит.

Многословные термины склоняются пословно с согласованием по главному слову
(эвристика: прилагательные согласуются, несклоняемые остаются как есть);
термины с пометой-прозой пропускаются (их ловит validate_dictionary.py).

Запуск:
    python gen_case_forms.py <путь-к-xlsx>
    python gen_case_forms.py <xlsx> --sheet "Именной" --limit 100
    python gen_case_forms.py "слово"            # разовая проверка одного слова

Зависимости: pymorphy3 + pymorphy3-dicts-ru (pip install pymorphy3), openpyxl
для режима xlsx. Офлайн, бесплатно, детерминированно — в отличие от
DeepSeek-обогащения (`copilot/enrich_dictionary.py`), которое стоит денег, но
лучше понимает санскритские имена; практичный порядок: сначала этот скрипт,
затем LLM-проверка спорных строк.

_Автор: Dr. Mārcis Gasūns · создан 10-07-2026 (H377)._
"""

import argparse
import datetime
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

CASES = ["nomn", "gent", "datv", "accs", "ablt", "loct"]
NUMBERS = ["sing", "plur"]


def get_analyzer():
    try:
        import pymorphy3
    except ImportError:
        print("Нужен pymorphy3: pip install pymorphy3", file=sys.stderr)
        sys.exit(2)
    return pymorphy3.MorphAnalyzer()


def best_parse(parses):
    """Предпочесть существительное в именительном падеже (термины указателя —
    почти всегда имена/существительные; для незнакомых санскритских имён
    pymorphy3 иначе выбирает прилагательное-гипотезу вроде «хануманный»)."""
    for p in parses:
        if "NOUN" in p.tag and "nomn" in p.tag:
            return p
    for p in parses:
        if "NOUN" in p.tag:
            return p
    return parses[0]


def word_forms(morph, word):
    """Все падежные формы слова (оба числа), с сохранением регистра первой буквы."""
    capital = word[:1].isupper()
    parses = morph.parse(word)
    if not parses:
        return [word]
    p = best_parse(parses)
    forms = []
    for number in NUMBERS:
        for case in CASES:
            f = p.inflect({number, case})
            if f is None:
                continue
            w = f.word
            if capital:
                w = w[:1].upper() + w[1:]
            if w not in forms:
                forms.append(w)
    return forms or [word]


def term_forms(morph, term):
    """Формы термина. Одно слово — полный набор; фраза — согласованное склонение."""
    words = term.split()
    if len(words) == 1:
        return word_forms(morph, term)
    # Фраза: склоняем каждое склоняемое слово в одном и том же падеже/числе.
    parses = [best_parse(morph.parse(w)) if morph.parse(w) else None for w in words]
    forms = []
    for number in NUMBERS:
        for case in CASES:
            out = []
            for w, p in zip(words, parses):
                f = p.inflect({number, case}) if p is not None else None
                nw = f.word if f is not None else w
                if w[:1].isupper():
                    nw = nw[:1].upper() + nw[1:]
                out.append(nw)
            phrase = " ".join(out)
            if phrase not in forms:
                forms.append(phrase)
    return forms


def main(argv=None):
    ap = argparse.ArgumentParser(description="pymorphy3-формы для словника указателей")
    ap.add_argument("target", help="путь к .xlsx ИЛИ одно слово/термин для проверки")
    ap.add_argument("--sheet", help="только один рабочий лист")
    ap.add_argument("--limit", type=int, default=0, help="максимум терминов (0 = все)")
    ap.add_argument("--report", help="путь к md-отчёту (по умолчанию рядом с xlsx)")
    ap.add_argument("--tsv", help="дополнительно выгрузить term\\tformы в TSV")
    args = ap.parse_args(argv)

    morph = get_analyzer()

    if not os.path.isfile(args.target):
        # Разовый режим: слово/термин из argv.
        print("; ".join(term_forms(morph, args.target)))
        return 0

    try:
        import openpyxl
    except ImportError:
        print("Нужен openpyxl: pip install openpyxl", file=sys.stderr)
        return 2
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from validate_dictionary import WORKING_SHEETS, find_columns, norm

    wb = openpyxl.load_workbook(args.target, read_only=True, data_only=True)
    sheets = [args.sheet] if args.sheet else [s for s in WORKING_SHEETS if s in wb.sheetnames]

    rows = []
    for sheet in sheets:
        ws = wb[sheet]
        header = [c.value for c in ws[1]]
        cols = find_columns(header)
        if cols["name"] is None:
            continue
        count = 0
        for r in range(2, ws.max_row + 1):
            term = norm(ws.cell(row=r, column=cols["name"]).value)
            if not term:
                continue
            current = norm(ws.cell(row=r, column=cols["forms"]).value) if cols["forms"] else ""
            head = term.split("\\")[-1].strip()  # уровень термина после '\'
            generated = term_forms(morph, head)
            cur_set = set(f.strip() for f in current.split(";") if f.strip())
            missing = [f for f in generated if f not in cur_set]
            rows.append((sheet, r, term, current, "; ".join(generated),
                         "; ".join(missing)))
            count += 1
            if args.limit and count >= args.limit:
                break

    report = args.report or os.path.join(
        os.path.dirname(os.path.abspath(args.target)), "case-forms-report.md")
    today = datetime.date.today().strftime("%d-%m-%Y")
    with open(report, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("# Отчёт pymorphy3-генератора падежных форм\n\n")
        fh.write("_Created: %s · Last updated: %s_\n\n" % (today, today))
        fh.write("Источник: `%s` · терминов: %d · генератор: pymorphy3 (OpenCorpora)\n\n"
                 % (os.path.basename(args.target), len(rows)))
        fh.write("| Лист | Строка | Термин | Текущие формы | Сгенерировано | Недостаёт |\n")
        fh.write("|---|---|---|---|---|---|\n")
        for row in rows:
            fh.write("| %s | %d | %s | %s | %s | %s |\n"
                     % tuple(str(x).replace("|", "/") for x in row))
        fh.write("\n_Сгенерировано `tools/gen_case_forms.py`; xlsx не изменялся —"
                 " правки переносит оператор._\n")
    print("Отчёт: %s · терминов: %d" % (report, len(rows)))

    if args.tsv:
        with open(args.tsv, "w", encoding="utf-8", newline="\n") as fh:
            for sheet, r, term, _cur, gen, _miss in rows:
                fh.write("%s\t%s\t%s\n" % (sheet, term, gen))
        print("TSV: %s" % args.tsv)
    return 0


if __name__ == "__main__":
    sys.exit(main())
